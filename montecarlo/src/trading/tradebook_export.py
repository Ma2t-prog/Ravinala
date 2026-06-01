"""
tradebook_export.py — Export the Ravinala Trade Book to Excel, PDF, CSV.
Dependencies: openpyxl (Excel), reportlab (PDF), csv (stdlib).
"""

import csv
import io
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from tradebook_models import (
    Trade, Book, PricingResult, PRODUCT_TYPE_LABELS,
    STATUS_ICONS, ASSET_CLASS_LABELS
)


# ═══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════

class TradeBookExporter:

    APP_NAME = "Ravinala v2.0"
    COPYRIGHT = "© 2026 TSIVAHINY Matthias — CONFIDENTIAL"

    # ─────────────────────────────────────────────────────────
    # EXCEL
    # ─────────────────────────────────────────────────────────

    def export_book_to_excel(self, book: Book, filepath: str = None,
                             book_metrics: dict = None) -> str:
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side,
                numbers as xl_numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl")

        if filepath is None:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filepath = str(Path(tempfile.gettempdir()) / f"ravinala_book_{ts}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        trades = book.get_trades()

        # ── Color palette ──────────────────────────────────
        NAVY = "0F172A"
        EMERALD = "059669"
        GOLD = "FBBF24"
        WHITE = "FFFFFF"
        LIGHT_GRAY = "F1F5F9"
        GREEN_FILL = "DCFCE7"
        RED_FILL = "FEE2E2"
        HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=11)
        HEADER_FILL = PatternFill("solid", fgColor=NAVY)
        SUBHEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
        ALT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
        thin = Side(style='thin', color='CBD5E1')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _header_row(ws, headers, row=1):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = BORDER

        def _set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def _accounting(n):
            if n is None:
                return ''
            return round(n, 2)

        # ── Sheet 1: Summary ───────────────────────────────
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        ws['A1'] = "RAVINALA -- Trade Book Summary"
        ws['A1'].font = Font(name='Calibri', bold=True, size=16, color=EMERALD)
        ws['A2'] = f"Book: {book.name}   |   Currency: {book.currency}   |   Generated: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"
        ws['A2'].font = Font(name='Calibri', size=10, color='64748B')
        ws.row_dimensions[1].height = 28
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')

        if book_metrics:
            kpis = [
                ("Total Trades", book_metrics.get('total_trades', 0), None),
                ("Live Trades", book_metrics.get('live_trades', 0), None),
                ("Total Notional", _accounting(book_metrics.get('total_notional')), book.currency),
                ("Total MTM", _accounting(book_metrics.get('total_mtm')), book.currency),
                ("Total P&L", _accounting(book_metrics.get('total_pnl')), book.currency),
                ("Agg. Delta", book_metrics.get('aggregate_delta', 0), ''),
                ("Agg. Vega", book_metrics.get('aggregate_vega', 0), ''),
                ("VaR 95% 1d", _accounting(book_metrics.get('portfolio_var_95')), book.currency),
            ]
            ws.cell(row=4, column=1, value="KEY METRICS").font = Font(bold=True, color=NAVY, size=12)
            for r, (label, val, ccy) in enumerate(kpis, 5):
                ws.cell(row=r, column=1, value=label).font = Font(bold=True)
                ws.cell(row=r, column=2, value=val)
                if ccy:
                    ws.cell(row=r, column=3, value=ccy).font = Font(color='64748B', italic=True)
                if r % 2 == 0:
                    for c in range(1, 4):
                        ws.cell(row=r, column=c).fill = ALT_FILL

        _set_col_widths(ws, [28, 18, 8, 8, 8, 8, 8])

        # ── Sheet 2: Deal Blotter ──────────────────────────
        ws2 = wb.create_sheet("Deal Blotter")
        ws2.freeze_panes = "A3"
        ws2.sheet_view.showGridLines = False

        headers2 = [
            "Ref", "Product Name", "Type", "Underlyings",
            "Dir", "Notional", "CCY", "Inception", "Maturity",
            "Strike %", "Entry Price %", "Current MTM", "P&L",
            "Δ Delta", "V Vega", "θ Theta", "Status", "Tags", "Desk"
        ]
        _header_row(ws2, headers2)
        ws2.row_dimensions[1].height = 22
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

        for r, t in enumerate(trades, 2):
            pr = t.get_current_pricing()
            row_vals = [
                t.internal_ref or t.trade_id[:8],
                t.product_name,
                PRODUCT_TYPE_LABELS.get(t.product_type, t.product_type),
                ', '.join(u.get('ticker', '') for u in t.underlyings),
                t.direction.upper(),
                _accounting(t.notional),
                t.currency,
                t.inception_date,
                t.maturity_date,
                t.strike_pct,
                t.entry_price,
                _accounting(t.current_mtm),
                _accounting(t.total_pnl),
                round(pr.delta, 4) if pr and pr.delta is not None else '',
                round(pr.vega, 4) if pr and pr.vega is not None else '',
                round(pr.theta, 4) if pr and pr.theta is not None else '',
                STATUS_ICONS.get(t.status, '') + ' ' + t.status.upper(),
                ', '.join(t.tags),
                t.desk,
            ]
            for c, v in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center')

            # Conditional formatting on P&L
            pnl_cell = ws2.cell(row=r, column=13)
            if isinstance(pnl_cell.value, (int, float)):
                if pnl_cell.value > 0:
                    pnl_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
                    pnl_cell.font = Font(color="16A34A", bold=True)
                elif pnl_cell.value < 0:
                    pnl_cell.fill = PatternFill("solid", fgColor=RED_FILL)
                    pnl_cell.font = Font(color="DC2626", bold=True)

            # Alternate rows
            if r % 2 == 0:
                for c in range(1, len(headers2) + 1):
                    if ws2.cell(row=r, column=c).fill.patternType != 'solid' or \
                       ws2.cell(row=r, column=c).fill.fgColor.rgb in (GREEN_FILL, RED_FILL, '00000000'):
                        if ws2.cell(row=r, column=c).fill.fgColor.rgb == '00000000':
                            ws2.cell(row=r, column=c).fill = ALT_FILL

        _set_col_widths(ws2, [14, 28, 20, 18, 5, 14, 5, 12, 12, 9, 12, 14, 14, 9, 9, 9, 14, 20, 14])

        # ── Sheet 3: Greeks ────────────────────────────────
        ws3 = wb.create_sheet("Greeks")
        ws3.freeze_panes = "A2"
        ws3.sheet_view.showGridLines = False
        headers3 = ["Ref", "Product", "Δ Delta", "Γ Gamma", "V Vega", "θ Theta", "ρ Rho", "Vanna", "Volga"]
        _header_row(ws3, headers3)

        agg = {k: 0.0 for k in ['delta', 'gamma', 'vega', 'theta', 'rho', 'vanna', 'volga']}
        for r, t in enumerate(trades, 2):
            pr = t.get_current_pricing()

            def g(name):
                val = getattr(pr, name, None) if pr else None
                if val is not None:
                    agg[name] = agg.get(name, 0) + val
                    return round(val, 6)
                return ''

            row_vals = [t.internal_ref, t.product_name,
                        g('delta'), g('gamma'), g('vega'),
                        g('theta'), g('rho'), g('vanna'), g('volga')]
            for c, v in enumerate(row_vals, 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.border = BORDER
                if r % 2 == 0:
                    cell.fill = ALT_FILL

        # Totals row
        tot_row = len(trades) + 2
        ws3.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
        ws3.cell(row=tot_row, column=1).fill = PatternFill("solid", fgColor=NAVY)
        ws3.cell(row=tot_row, column=2, value="Aggregate").font = Font(bold=True, color=WHITE)
        ws3.cell(row=tot_row, column=2).fill = PatternFill("solid", fgColor=NAVY)
        for c, key in enumerate(['delta', 'gamma', 'vega', 'theta', 'rho', 'vanna', 'volga'], 3):
            cell = ws3.cell(row=tot_row, column=c, value=round(agg[key], 6))
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.border = BORDER

        _set_col_widths(ws3, [14, 28, 12, 12, 12, 12, 12, 12, 12])

        # ── Sheet 4: P&L Attribution ───────────────────────
        ws4 = wb.create_sheet("P&L Attribution")
        ws4.sheet_view.showGridLines = False
        ws4['A1'] = "P&L Attribution by Trade"
        ws4['A1'].font = Font(bold=True, size=13, color=NAVY)
        headers4 = ["Ref", "Product", "Notional", "Entry Price %", "Current MTM", "Unrealized P&L", "Realized P&L", "Total P&L"]
        _header_row(ws4, headers4, row=2)
        for r, t in enumerate(trades, 3):
            row_vals = [t.internal_ref, t.product_name,
                        _accounting(t.notional), t.entry_price,
                        _accounting(t.current_mtm),
                        _accounting(t.unrealized_pnl),
                        _accounting(t.realized_pnl),
                        _accounting(t.total_pnl)]
            for c, v in enumerate(row_vals, 1):
                cell = ws4.cell(row=r, column=c, value=v)
                cell.border = BORDER
                if c in (6, 7, 8) and isinstance(v, (int, float)):
                    cell.fill = PatternFill("solid",
                                            fgColor=GREEN_FILL if v >= 0 else RED_FILL)
                    cell.font = Font(color="16A34A" if v >= 0 else "DC2626", bold=True)

        _set_col_widths(ws4, [14, 28, 14, 14, 14, 16, 16, 16])

        # ── Sheet 5: Maturity Profile ──────────────────────
        ws5 = wb.create_sheet("Maturity Profile")
        ws5.sheet_view.showGridLines = False
        ws5['A1'] = "Maturity Schedule"
        ws5['A1'].font = Font(bold=True, size=13, color=NAVY)
        headers5 = ["Ref", "Product", "Underlying", "Maturity", "Notional", "CCY", "Status"]
        _header_row(ws5, headers5, row=2)
        sorted_trades = sorted(trades, key=lambda t: t.maturity_date or '')
        for r, t in enumerate(sorted_trades, 3):
            tickers = ', '.join(u.get('ticker', '') for u in t.underlyings)
            row_vals = [t.internal_ref, t.product_name, tickers,
                        t.maturity_date, _accounting(t.notional),
                        t.currency, t.status.upper()]
            for c, v in enumerate(row_vals, 1):
                cell = ws5.cell(row=r, column=c, value=v)
                cell.border = BORDER
                if r % 2 == 0:
                    cell.fill = ALT_FILL

        _set_col_widths(ws5, [14, 28, 16, 14, 14, 6, 12])

        # ── Footer on all sheets ───────────────────────────
        footer_text = f"&L{self.APP_NAME}&R{self.COPYRIGHT}"
        for ws_obj in wb.worksheets:
            ws_obj.oddFooter.center.text = footer_text
            ws_obj.oddFooter.center.size = 8

        wb.save(filepath)
        return filepath

    # ─────────────────────────────────────────────────────────
    # CSV
    # ─────────────────────────────────────────────────────────

    def export_blotter_to_csv(self, book: Book, filepath: str = None) -> str:
        if filepath is None:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filepath = str(Path(tempfile.gettempdir()) / f"ravinala_blotter_{ts}.csv")

        trades = book.get_trades()
        fieldnames = [
            "ref", "product_name", "product_type", "underlyings",
            "direction", "notional", "currency", "inception_date", "maturity_date",
            "tenor_years", "strike_pct", "entry_price", "current_mtm",
            "unrealized_pnl", "realized_pnl", "total_pnl",
            "delta", "vega", "theta", "status", "counterparty", "desk", "tags", "notes"
        ]

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for t in trades:
                pr = t.get_current_pricing()
                writer.writerow({
                    'ref': t.internal_ref,
                    'product_name': t.product_name,
                    'product_type': t.product_type,
                    'underlyings': '|'.join(u.get('ticker', '') for u in t.underlyings),
                    'direction': t.direction,
                    'notional': t.notional,
                    'currency': t.currency,
                    'inception_date': t.inception_date,
                    'maturity_date': t.maturity_date,
                    'tenor_years': t.tenor_years,
                    'strike_pct': t.strike_pct,
                    'entry_price': t.entry_price,
                    'current_mtm': t.current_mtm,
                    'unrealized_pnl': t.unrealized_pnl,
                    'realized_pnl': t.realized_pnl,
                    'total_pnl': t.total_pnl,
                    'delta': pr.delta if pr else '',
                    'vega': pr.vega if pr else '',
                    'theta': pr.theta if pr else '',
                    'status': t.status,
                    'counterparty': t.counterparty,
                    'desk': t.desk,
                    'tags': ', '.join(t.tags),
                    'notes': t.notes,
                })

        return filepath

    # ─────────────────────────────────────────────────────────
    # PDF — Internal Term Sheet
    # ─────────────────────────────────────────────────────────

    def export_trade_to_pdf(self, trade: Trade, filepath: str = None) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                HRFlowable
            )
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        except ImportError:
            raise ImportError("reportlab required: pip install reportlab")

        if filepath is None:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            ref_safe = (trade.internal_ref or trade.trade_id).replace('/', '_')
            filepath = str(Path(tempfile.gettempdir()) / f"termsheet_{ref_safe}_{ts}.pdf")

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            topMargin=1.5*cm, bottomMargin=2*cm,
            leftMargin=2*cm, rightMargin=2*cm
        )

        # ── Color palette ──
        NAVY = colors.HexColor('#0F172A')
        EMERALD = colors.HexColor('#059669')
        TEAL = colors.HexColor('#2DD4BF')
        GOLD = colors.HexColor('#FBBF24')
        LIGHT_BG = colors.HexColor('#F1F5F9')
        GREEN = colors.HexColor('#16A34A')
        RED = colors.HexColor('#DC2626')
        MID_GRAY = colors.HexColor('#64748B')

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                     textColor=NAVY, fontSize=16, spaceAfter=4, spaceBefore=0)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'],
                                       textColor=EMERALD, fontSize=10, spaceAfter=4,
                                       spaceBefore=8, fontName='Helvetica-Bold')
        body_style = ParagraphStyle('Body', parent=styles['Normal'],
                                    fontSize=9, textColor=NAVY, leading=14)
        small_style = ParagraphStyle('Small', parent=styles['Normal'],
                                     fontSize=8, textColor=MID_GRAY, leading=12)
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
                                      fontSize=7, textColor=MID_GRAY,
                                      alignment=TA_CENTER)

        # ── Table style helper ──
        def kv_table(rows, col_widths=None):
            if col_widths is None:
                col_widths = [6*cm, 11*cm]
            t = Table(rows, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, 0), (0, -1), NAVY),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1E3A5F')),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1),
                 [colors.white, LIGHT_BG]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            return t

        story = []

        # ── HEADER ──
        story.append(Paragraph(f"{self.APP_NAME} -- Internal Term Sheet", title_style))
        story.append(HRFlowable(width="100%", thickness=2, color=EMERALD, spaceAfter=8))

        # ── IDENTITY ──
        story.append(Paragraph("TRADE IDENTITY", section_style))
        story.append(kv_table([
            ["Reference", trade.internal_ref or trade.trade_id],
            ["External Ref", trade.external_ref or "—"],
            ["Product", trade.product_name or PRODUCT_TYPE_LABELS.get(trade.product_type, trade.product_type)],
            ["Status", STATUS_ICONS.get(trade.status, '') + " " + trade.status.upper()],
            ["Version", f"V{trade.current_version}"],
        ]))
        story.append(Spacer(1, 0.3*cm))

        # ── TRADE DETAILS ──
        story.append(Paragraph("TRADE DETAILS", section_style))
        story.append(kv_table([
            ["Counterparty", trade.counterparty or "—"],
            ["Desk", trade.desk or "—"],
            ["Direction", trade.direction.upper()],
            ["Notional", f"{trade.currency} {trade.notional:,.0f}"],
            ["Trade Date", trade.trade_date or "—"],
            ["Inception Date", trade.inception_date or "—"],
            ["Maturity Date", trade.maturity_date or "—"],
            ["Tenor", f"{trade.tenor_years:.2f} years" if trade.tenor_years else "—"],
        ]))
        story.append(Spacer(1, 0.3*cm))

        # ── UNDERLYINGS ──
        story.append(Paragraph("UNDERLYING(S)", section_style))
        ul_rows = [["Ticker", "Name", "Asset Class", "Spot @ Inception", "Current Spot", "Perf %"]]
        for u in trade.underlyings:
            s0 = u.get('spot_at_inception', 0)
            sc = u.get('current_spot') or s0
            perf = ((sc - s0) / s0 * 100) if s0 else None
            perf_str = f"{perf:+.2f}%" if perf is not None else "—"
            ul_rows.append([
                u.get('ticker', ''),
                u.get('name', ''),
                ASSET_CLASS_LABELS.get(u.get('asset_class', ''), u.get('asset_class', '')),
                f"{s0:,.2f}" if s0 else "—",
                f"{sc:,.2f}" if sc else "—",
                perf_str,
            ])
        ul_t = Table(ul_rows, colWidths=[3*cm, 4*cm, 3*cm, 3*cm, 3*cm, 2*cm])
        ul_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(ul_t)
        story.append(Spacer(1, 0.3*cm))

        # ── STRUCTURE ──
        story.append(Paragraph("STRUCTURE", section_style))
        struct_rows = [["Product Type", PRODUCT_TYPE_LABELS.get(trade.product_type, trade.product_type)]]
        if trade.strike_pct is not None:
            struct_rows.append(["Strike", f"{trade.strike_pct:.1f}% of spot at inception"])
        if trade.capital_protection_pct is not None:
            struct_rows.append(["Capital Protection", f"{trade.capital_protection_pct:.1f}%"])
        if trade.participation_rate is not None:
            struct_rows.append(["Participation Rate", f"{trade.participation_rate:.1f}%"])
        if trade.cap_pct is not None:
            struct_rows.append(["Cap", f"{trade.cap_pct:.1f}%"])

        for b in trade.barriers:
            btype = b.get('barrier_type', '').replace('_', ' ').title()
            obs = b.get('observation', '').replace('_', ' ')
            struct_rows.append([
                f"Barrier ({btype})",
                f"{b.get('level_pct', '')}% — {obs}",
            ])

        if trade.coupon:
            c = trade.coupon
            struct_rows.append(["Coupon Rate", f"{c.get('rate_pct', '')}% {c.get('frequency', '')}"])
            struct_rows.append(["Coupon Memory", "Yes" if c.get('is_memory') else "No"])
            if c.get('condition_barrier_pct'):
                struct_rows.append(["Coupon Barrier", f"{c['condition_barrier_pct']}% of inception spot"])

        story.append(kv_table(struct_rows))
        story.append(Spacer(1, 0.3*cm))

        # ── PRICING ──
        pr = trade.get_current_pricing()
        if pr:
            story.append(Paragraph("PRICING", section_style))
            stale_note = " STALE" if pr.is_stale else ""
            story.append(kv_table([
                ["Model", pr.model.replace('_', ' ').title() + stale_note],
                ["Price", f"{pr.price:.4f}%" if pr.price is not None else "—"],
                ["MTM Value", f"{trade.currency} {pr.notional_value:,.0f}" if pr.notional_value else "—"],
                ["Implied Vol", f"{pr.vol_used*100:.2f}%" if pr.vol_used else "—"],
                ["Risk-Free Rate", f"{pr.rate_used*100:.2f}%" if pr.rate_used else "—"],
                ["Pricing Date", pr.timestamp[:10] if pr.timestamp else "—"],
                ["MC Paths", f"{pr.mc_paths:,}" if pr.mc_paths else "—"],
                ["Std Error", f"±{pr.mc_std_error:.4f}%" if pr.mc_std_error else "—"],
            ]))
            story.append(Spacer(1, 0.2*cm))

            # Greeks table
            story.append(Paragraph("GREEKS", section_style))
            greek_rows = []
            for name, val in [
                ("Delta (Δ)", pr.delta), ("Gamma (Γ)", pr.gamma),
                ("Vega (V)", pr.vega), ("Theta (θ)", pr.theta),
                ("Rho (ρ)", pr.rho), ("Vanna", pr.vanna), ("Volga", pr.volga),
            ]:
                if val is not None:
                    greek_rows.append([name, f"{val:.6f}"])
            if greek_rows:
                story.append(kv_table(greek_rows, col_widths=[5*cm, 5*cm]))
            story.append(Spacer(1, 0.3*cm))

        # ── P&L ──
        story.append(Paragraph("P&L", section_style))
        def pnl_str(v):
            if v is None:
                return "—"
            sign = '+' if v >= 0 else ''
            return f"{sign}{v:,.2f} {trade.pnl_currency}"

        story.append(kv_table([
            ["Entry Price", f"{trade.entry_price:.4f}%" if trade.entry_price else "—"],
            ["Current MTM", f"{trade.currency} {trade.current_mtm:,.0f}" if trade.current_mtm else "—"],
            ["Unrealized P&L", pnl_str(trade.unrealized_pnl)],
            ["Realized P&L", pnl_str(trade.realized_pnl)],
            ["Total P&L", pnl_str(trade.total_pnl)],
        ]))
        story.append(Spacer(1, 0.3*cm))

        # ── NOTES ──
        if trade.notes:
            story.append(Paragraph("NOTES", section_style))
            story.append(Paragraph(trade.notes, body_style))
            story.append(Spacer(1, 0.3*cm))

        # ── TAGS ──
        if trade.tags:
            story.append(Paragraph("TAGS", section_style))
            story.append(Paragraph("  ".join(f"[{tag}]" for tag in trade.tags), small_style))
            story.append(Spacer(1, 0.3*cm))

        # ── AUDIT TRAIL ──
        if trade.audit_trail:
            story.append(Paragraph("AUDIT TRAIL", section_style))
            audit_rows = [["Date", "User", "Action", "Details"]]
            for entry in reversed(trade.audit_trail[-10:]):
                audit_rows.append([
                    entry.get('timestamp', '')[:16],
                    entry.get('user', ''),
                    entry.get('action', ''),
                    entry.get('details', '')[:60],
                ])
            at = Table(audit_rows, colWidths=[4*cm, 3*cm, 3.5*cm, 7*cm])
            at.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(at)
            story.append(Spacer(1, 0.5*cm))

        # ── FOOTER ──
        story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f"Generated by {self.APP_NAME}  |  "
            f"{datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')}  |  "
            f"CONFIDENTIAL — For internal use only",
            footer_style
        ))

        doc.build(story)
        return filepath

    # ─────────────────────────────────────────────────────────
    # RISK REPORT (PDF)
    # ─────────────────────────────────────────────────────────

    def export_risk_report(self, book: Book, book_metrics: dict,
                           filepath: str = None) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                HRFlowable
            )
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            raise ImportError("reportlab required: pip install reportlab")

        if filepath is None:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filepath = str(Path(tempfile.gettempdir()) / f"ravinala_risk_report_{ts}.pdf")

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                topMargin=1.5*cm, bottomMargin=2*cm,
                                leftMargin=2*cm, rightMargin=2*cm)

        NAVY = colors.HexColor('#0F172A')
        EMERALD = colors.HexColor('#059669')
        LIGHT_BG = colors.HexColor('#F1F5F9')
        MID_GRAY = colors.HexColor('#64748B')

        styles = getSampleStyleSheet()
        title_s = ParagraphStyle('T', parent=styles['Heading1'],
                                  textColor=NAVY, fontSize=16, spaceAfter=4)
        sec_s = ParagraphStyle('S', parent=styles['Heading2'],
                               textColor=EMERALD, fontSize=11, spaceAfter=4, spaceBefore=10)
        body_s = ParagraphStyle('B', parent=styles['Normal'], fontSize=9)

        story = []
        story.append(Paragraph(f"{self.APP_NAME} -- Risk Report", title_s))
        story.append(Paragraph(
            f"Book: {book.name}  |  {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}",
            ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, textColor=MID_GRAY)
        ))
        story.append(HRFlowable(width="100%", thickness=2, color=EMERALD, spaceAfter=10))

        # Risk overview
        story.append(Paragraph("PORTFOLIO RISK OVERVIEW", sec_s))
        risk_rows = [
            ["Metric", "Value", "Currency"],
            ["Total Notional", f"{book_metrics.get('total_notional', 0):,.0f}", book.currency],
            ["VaR 95% 1d", f"{book_metrics.get('portfolio_var_95', 0):,.0f}", book.currency],
            ["Live Trades", str(book_metrics.get('live_trades', 0)), ""],
            ["Aggregate Delta", f"{book_metrics.get('aggregate_delta', 0):.4f}", ""],
            ["Aggregate Vega", f"{book_metrics.get('aggregate_vega', 0):.4f}", ""],
            ["Aggregate Theta", f"{book_metrics.get('aggregate_theta', 0):.4f}", ""],
        ]
        t = Table(risk_rows, colWidths=[8*cm, 6*cm, 4*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)

        # Concentration
        conc = book_metrics.get('concentration', {})
        if conc:
            story.append(Paragraph("TOP CONCENTRATIONS BY UNDERLYING", sec_s))
            conc_rows = [["Underlying", f"Notional ({book.currency})", "% of Book"]]
            total_not = book_metrics.get('total_notional', 1) or 1
            for tk, val in conc.items():
                conc_rows.append([tk, f"{val:,.0f}", f"{val/total_not*100:.1f}%"])
            ct = Table(conc_rows, colWidths=[8*cm, 6*cm, 4*cm])
            ct.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), NAVY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
                ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(ct)

        # Footer
        story.append(Spacer(1, 1*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY))
        story.append(Paragraph(
            f"Generated by {self.APP_NAME}  |  {self.COPYRIGHT}",
            ParagraphStyle('f', parent=styles['Normal'], fontSize=7,
                           textColor=MID_GRAY, alignment=TA_CENTER)
        ))

        doc.build(story)
        return filepath

    # ─────────────────────────────────────────────────────────
    # IN-MEMORY HELPERS (for Streamlit download buttons)
    # ─────────────────────────────────────────────────────────

    def export_csv_bytes(self, book: Book) -> bytes:
        """Returns CSV as bytes (for st.download_button)."""
        buf = io.StringIO()
        trades = book.get_trades()
        fieldnames = [
            "ref", "product_name", "product_type", "underlyings",
            "direction", "notional", "currency", "inception_date", "maturity_date",
            "entry_price", "current_mtm", "total_pnl", "status", "tags"
        ]
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for t in trades:
            writer.writerow({
                'ref': t.internal_ref, 'product_name': t.product_name,
                'product_type': t.product_type,
                'underlyings': '|'.join(u.get('ticker', '') for u in t.underlyings),
                'direction': t.direction, 'notional': t.notional, 'currency': t.currency,
                'inception_date': t.inception_date, 'maturity_date': t.maturity_date,
                'entry_price': t.entry_price, 'current_mtm': t.current_mtm,
                'total_pnl': t.total_pnl, 'status': t.status, 'tags': ', '.join(t.tags),
            })
        return buf.getvalue().encode('utf-8')

    def export_excel_bytes(self, book: Book, book_metrics: dict = None) -> bytes:
        """Returns Excel as bytes (for st.download_button)."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            tmp = f.name
        try:
            self.export_book_to_excel(book, tmp, book_metrics)
            with open(tmp, 'rb') as f:
                return f.read()
        finally:
            try:
                os.unlink(tmp)
            except Exception:
                pass

    def export_pdf_bytes(self, trade: Trade) -> bytes:
        """Returns PDF as bytes (for st.download_button)."""
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
            tmp = f.name
        try:
            self.export_trade_to_pdf(trade, tmp)
            with open(tmp, 'rb') as f:
                return f.read()
        finally:
            try:
                os.unlink(tmp)
            except Exception:
                pass
