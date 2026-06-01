"""Export GenesiX analysis results to Excel workbooks."""

from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(results: dict, portfolio_weights: dict, investment: float) -> bytes:
    """
    Generate Excel workbook with GenesiX analysis results.
    
    Multi-sheet workbook:
    - Summary: Key metrics
    - Scenarios: Probability-weighted outcomes
    - Risk Metrics: VaR, CVaR, volatility
    - Portfolio: Weights and allocations
    - Raw Data: JSON export of full results
    
    Args:
        results: Portfolio analysis results dict
        portfolio_weights: {asset: weight} dict
        investment: Investment amount (EUR)
    
    Returns:
        XLSX as bytes
    """
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    subheader_fill = PatternFill(start_color='D0D0E8', end_color='D0D0E8', fill_type='solid')
    positive_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    negative_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Build sheets
    _add_summary_sheet(wb, results, investment, header_font, header_fill, subheader_fill, border)
    _add_scenarios_sheet(wb, results, investment, header_font, header_fill, positive_fill, negative_fill, border)
    _add_risk_metrics_sheet(wb, results, investment, header_font, header_fill, border)
    _add_portfolio_sheet(wb, portfolio_weights, investment, header_font, header_fill, border)
    _add_raw_data_sheet(wb, results, header_font, header_fill, border)
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _add_summary_sheet(wb, results, investment, header_font, header_fill, subheader_fill, border):
    """Create summary sheet."""
    ws = wb.create_sheet('Summary', 0)
    
    # Title
    ws['A1'] = 'GenesiX Portfolio Analysis'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    # Key metrics
    row = 4
    ws[f'A{row}'] = 'Key Metrics'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = subheader_fill
    row += 1
    
    pred = results.get('prediction', {})
    rm = results.get('risk_metrics', {})
    
    metrics = [
        ('Investment Amount', f"€{investment:,.2f}"),
        ('Expected Return', f"{pred.get('expected_return_pct', 0):+.2f}%"),
        ('Expected Value', f"€{investment * (1 + pred.get('expected_return_pct', 0)/100):,.2f}"),
        ('Probability of Profit', f"{pred.get('probability_positive', 0):.1%}"),
        ('5th Percentile (Worst)', f"€{investment * (1 + pred.get('worst_case_pct', 0)/100):,.2f}"),
        ('95th Percentile (Best)', f"€{investment * (1 + pred.get('best_case_pct', 0)/100):,.2f}"),
        ('', ''),
        ('Annual Volatility', f"{rm.get('volatility_annualized', 0):.1%}"),
        ('VaR 95% (1-day)', f"{rm.get('var_95', 0):.2%}"),
        ('CVaR 95% (Expected Shortfall)', f"{rm.get('cvar_95', 0):.2%}"),
        ('Maximum Drawdown', f"{rm.get('max_drawdown', 0):.1%}"),
        ('Sharpe Ratio', f"{rm.get('sharpe_ratio', 0):.2f}"),
        ('Diversification Ratio', f"{rm.get('diversification_ratio', 1):.2f}"),
    ]
    
    for label, value in metrics:
        if label == '':
            row += 1
        else:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def _add_scenarios_sheet(wb, results, investment, header_font, header_fill, positive_fill, negative_fill, border):
    """Create scenarios sheet."""
    ws = wb.create_sheet('Scenarios', 1)
    
    # Header
    ws['A1'] = 'Scenario Analysis'
    ws['A1'].font = Font(bold=True, size=12)
    
    # Table header
    headers = ['Scenario', 'Probability', 'Return %', 'Final Value (€)', 'P&L (€)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    scenarios = results.get('scenarios', [])
    row = 4
    for scenario in scenarios:
        name = scenario.get('name', 'Unknown')
        prob = scenario.get('probability', 0)
        ret = scenario.get('return_pct', 0)
        final = scenario.get('final_value', 0)
        pnl = final - investment
        
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).value = prob
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = ret / 100
        ws.cell(row=row, column=3).number_format = '0.00%'
        ws.cell(row=row, column=4).value = final
        ws.cell(row=row, column=4).number_format = '€#,##0.00'
        ws.cell(row=row, column=5).value = pnl
        ws.cell(row=row, column=5).number_format = '€#,##0.00'
        
        # Color rows
        fill = positive_fill if pnl >= 0 else negative_fill
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
        
        row += 1
    
    # Adjust widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18


def _add_risk_metrics_sheet(wb, results, investment, header_font, header_fill, border):
    """Create risk metrics sheet."""
    ws = wb.create_sheet('Risk Metrics', 2)
    
    ws['A1'] = 'Risk Analysis'
    ws['A1'].font = Font(bold=True, size=12)
    
    rm = results.get('risk_metrics', {})
    
    # VaR section
    row = 3
    headers = ['Horizon', 'VaR 95% (%)', 'VaR 95% (€)', 'CVaR 95% (%)', 'CVaR 95% (€)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    row = 4
    horizons = [
        ('1-day', rm.get('var_95', 0), rm.get('cvar_95', 0)),
        ('5-day', rm.get('var_95_5d', 0), rm.get('cvar_95_5d', 0)),
        ('21-day', rm.get('var_95_21d', 0), rm.get('cvar_95_21d', 0)),
    ]
    
    for horizon_label, var, cvar in horizons:
        ws.cell(row=row, column=1).value = horizon_label
        ws.cell(row=row, column=2).value = var
        ws.cell(row=row, column=2).number_format = '0.00%'
        ws.cell(row=row, column=3).value = investment * var
        ws.cell(row=row, column=3).number_format = '€#,##0.00'
        ws.cell(row=row, column=4).value = cvar
        ws.cell(row=row, column=4).number_format = '0.00%'
        ws.cell(row=row, column=5).value = investment * cvar
        ws.cell(row=row, column=5).number_format = '€#,##0.00'
        row += 1
    
    # Volatility and other metrics
    row += 2
    metrics = [
        ('Annual Volatility', rm.get('volatility_annualized', 0)),
        ('Maximum Drawdown', rm.get('max_drawdown', 0)),
        ('Sharpe Ratio', rm.get('sharpe_ratio', 0)),
        ('Sortino Ratio', rm.get('sortino_ratio', 0)),
        ('Calmar Ratio', rm.get('calmar_ratio', 0)),
        ('Diversification Ratio', rm.get('diversification_ratio', 1)),
    ]
    
    for label, value in metrics:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        if label != 'Sharpe Ratio' and label != 'Sortino Ratio' and label != 'Calmar Ratio':
            ws.cell(row=row, column=2).number_format = '0.00%'
        else:
            ws.cell(row=row, column=2).number_format = '0.00'
        row += 1
    
    # Adjust widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18


def _add_portfolio_sheet(wb, portfolio_weights, investment, header_font, header_fill, border):
    """Create portfolio composition sheet."""
    ws = wb.create_sheet('Portfolio', 3)
    
    ws['A1'] = 'Portfolio Composition'
    ws['A1'].font = Font(bold=True, size=12)
    
    # Header
    headers = ['Asset', 'Weight', 'Amount (€)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Data
    row = 4
    total = 0
    for asset, weight in sorted(portfolio_weights.items(), key=lambda x: x[1], reverse=True):
        amount = investment * weight
        total += amount
        
        ws.cell(row=row, column=1).value = asset
        ws.cell(row=row, column=2).value = weight
        ws.cell(row=row, column=2).number_format = '0.00%'
        ws.cell(row=row, column=3).value = amount
        ws.cell(row=row, column=3).number_format = '€#,##0.00'
        row += 1
    
    # Total row
    ws.cell(row=row, column=1).value = 'TOTAL'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = 1.0
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).value = total
    ws.cell(row=row, column=3).number_format = '€#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    
    # Adjust widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18


def _add_raw_data_sheet(wb, results, header_font, header_fill, border):
    """Create sheet with raw data export (JSON representation)."""
    ws = wb.create_sheet('Raw Data', 4)
    
    ws['A1'] = 'Raw Analysis Data (JSON Format)'
    ws['A1'].font = Font(bold=True, size=11)
    ws['A2'] = 'This sheet contains the complete results dict exported as text.'
    ws['A2'].font = Font(italic=True, size=9)
    
    # Export key sections
    row = 4
    sections = [
        ('prediction', results.get('prediction', {})),
        ('risk_metrics', results.get('risk_metrics', {})),
        ('scenarios', results.get('scenarios', [])),
        ('alert_level', results.get('alert_level', {})),
    ]
    
    import json
    for section_name, section_data in sections:
        ws.cell(row=row, column=1).value = section_name
        ws.cell(row=row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = header_fill
        row += 1
        
        json_str = json.dumps(section_data, indent=2, default=str)
        ws.cell(row=row, column=1).value = json_str
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        row += len(json_str.split('\n')) + 2
    
    ws.column_dimensions['A'].width = 80


__all__ = ['export_to_excel']
