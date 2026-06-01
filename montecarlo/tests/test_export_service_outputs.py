from __future__ import annotations

import sys
import uuid
from pathlib import Path

import openpyxl
import pytest

BACKEND_DIR = Path(__file__).resolve().parents[1] / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.models import ExcelExportRequest, PDFExportRequest
from app.services import export_service


def _sample_snapshot() -> dict:
    return {
        "indices": {
            "americas": [
                {
                    "symbol": "SPX",
                    "name": "S&P 500",
                    "region": "Americas",
                    "price": 5100.0,
                    "change": {"percent": 0.8},
                    "timestamp": "2026-03-24T00:00:00+00:00",
                }
            ]
        },
        "bonds": {
            "bonds": [
                {
                    "country": "United States",
                    "yield_2y": 4.5,
                    "yield_5y": 4.1,
                    "yield_10y": 3.9,
                    "spread_vs_bund_bp": 180,
                }
            ]
        },
        "timestamp": "2026-03-24T00:00:00+00:00",
        "cache_hit": False,
    }


@pytest.mark.asyncio
async def test_export_dashboard_excel_creates_readable_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def _snapshot():
        return _sample_snapshot()

    export_root = Path(__file__).resolve().parents[1] / "tmp" / f"export-test-{uuid.uuid4().hex}"
    export_root.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(export_service, "get_full_snapshot_async", _snapshot)
    monkeypatch.setattr(export_service.tempfile, "gettempdir", lambda: str(export_root))

    exported = await export_service.export_dashboard_excel(
        ExcelExportRequest(sheets=["indices", "bonds"])
    )

    path = Path(exported.path)
    assert path.exists()
    assert exported.filename.endswith(".xlsx")

    workbook = openpyxl.load_workbook(path)
    assert "Indices" in workbook.sheetnames
    assert "Bonds" in workbook.sheetnames
    assert workbook["Indices"]["A2"].value == "SPX"
    assert workbook["Bonds"]["A2"].value == "United States"


@pytest.mark.asyncio
async def test_export_dashboard_pdf_creates_pdf_file(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def _snapshot():
        return _sample_snapshot()

    export_root = Path(__file__).resolve().parents[1] / "tmp" / f"export-test-{uuid.uuid4().hex}"
    export_root.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(export_service, "get_full_snapshot_async", _snapshot)
    monkeypatch.setattr(export_service.tempfile, "gettempdir", lambda: str(export_root))

    exported = await export_service.export_dashboard_pdf(PDFExportRequest())

    path = Path(exported.path)
    assert path.exists()
    assert exported.filename.endswith(".pdf")
    assert path.read_bytes()[:4] == b"%PDF"
