"""
Tests for Global Macro Terminal — macro_data.py + macro_export.py
Run: pytest tests/test_macro.py -v
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from unittest.mock import patch, MagicMock, PropertyMock
import sys

# Stub streamlit before import
st_mock = MagicMock()
st_mock.cache_data = lambda **kw: (lambda f: f)
sys.modules['streamlit'] = st_mock

import pytest
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
import io


# ─────────────────────────────────────────────────────────────────
# HELPERS & FIXTURES
# ─────────────────────────────────────────────────────────────────

def _fake_closes(tickers: list, n: int = 35, base: float = 100.0) -> pd.DataFrame:
    """Build a synthetic close-price DataFrame for mocking yfinance download."""
    rng = np.random.default_rng(42)
    idx = pd.date_range(end='2026-03-12', periods=n, freq='B')
    data = {}
    for t in tickers:
        prices = base + rng.normal(0, 1, n).cumsum()
        prices = np.maximum(prices, 1.0)
        data[t] = prices
    return pd.DataFrame(data, index=idx)


def _fake_ticker_news():
    return [
        {'title': 'Fed holds rates steady',
         'publisher': 'Reuters',
         'link': 'https://reuters.com/1',
         'providerPublishTime': int(datetime.now(timezone.utc).timestamp()) - 3600},
        {'title': 'Gold hits all-time high',
         'publisher': 'Bloomberg',
         'link': 'https://bloomberg.com/2',
         'providerPublishTime': int(datetime.now(timezone.utc).timestamp()) - 7200},
    ]


@pytest.fixture(scope='module')
def provider():
    from macro_data import MacroDataProvider
    return MacroDataProvider()


# ─────────────────────────────────────────────────────────────────
# _spark / formatting helpers
# ─────────────────────────────────────────────────────────────────

class TestHelpers:
    def test_spark_returns_img_tag(self):
        from macro_data import _spark
        result = _spark([100, 101, 99, 102, 103])
        assert result.startswith('<img src="data:image/svg+xml;base64,')

    def test_spark_empty_returns_empty(self):
        from macro_data import _spark
        assert _spark([]) == ""
        assert _spark([100]) == ""

    def test_spark_flat_prices(self):
        from macro_data import _spark
        # All same value — should handle gracefully
        result = _spark([100.0, 100.0, 100.0])
        assert isinstance(result, str)

    def test_pct_span_positive(self):
        from macro_data import _pct_span
        html = _pct_span(1.5)
        assert 'color:#00C853' in html
        assert '+1.50%' in html

    def test_pct_span_negative(self):
        from macro_data import _pct_span
        html = _pct_span(-0.75)
        assert 'color:#FF1744' in html
        assert '-0.75%' in html

    def test_pct_span_none(self):
        from macro_data import _pct_span
        html = _pct_span(None)
        assert 'N/A' in html

    def test_fmt_large_number(self):
        from macro_data import _fmt
        assert ',' in _fmt(6750.0, thousands=True)

    def test_fmt_none(self):
        from macro_data import _fmt
        assert _fmt(None) == 'N/A'


# ─────────────────────────────────────────────────────────────────
# MacroDataProvider — unit tests with mocked yfinance
# ─────────────────────────────────────────────────────────────────

class TestMacroDataProvider:

    def _patch_download(self, provider, tickers, n=35):
        closes = _fake_closes(tickers, n)
        mock_df = pd.DataFrame()
        # Simulate MultiIndex columns as yfinance returns
        mi = pd.MultiIndex.from_tuples([('Close', t) for t in tickers])
        mock_df_mi = _fake_closes(tickers, n)
        mock_df_mi.columns = pd.MultiIndex.from_tuples([('Close', t) for t in tickers])
        return mock_df_mi

    @patch('yfinance.download')
    def test_fetch_equity_indices_returns_df(self, mock_dl, provider):
        tickers = list(provider.EQUITY_INDICES.keys())[:5]
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_equity_indices()
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.download')
    def test_fetch_equity_indices_columns(self, mock_dl, provider):
        tickers = list(provider.EQUITY_INDICES.keys())[:4]
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_equity_indices()
        if not df.empty:
            for col in ('name', 'flag', 'group', 'last', 'change_pct'):
                assert col in df.columns

    @patch('yfinance.download')
    def test_fetch_fx_returns_df(self, mock_dl, provider):
        tickers = list(provider.FX_G10.keys())[:4]
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_fx()
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.download')
    def test_fetch_fx_groups(self, mock_dl, provider):
        all_fx = {**provider.FX_G10, **provider.FX_EM}
        tickers = list(all_fx.keys())[:6]
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_fx()
        if not df.empty and 'group' in df.columns:
            groups = df['group'].unique()
            assert set(groups).issubset({'G10', 'EM'})

    @patch('yfinance.download')
    def test_fetch_commodities_returns_df(self, mock_dl, provider):
        tickers = list(provider.COMMODITIES.keys())[:5]
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_commodities()
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.download')
    def test_fetch_bond_yields_returns_df(self, mock_dl, provider):
        tickers = list(provider.BOND_TICKERS.keys())
        mock_dl.return_value = self._patch_download(provider, tickers, n=300)
        df = provider.fetch_bond_yields()
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.download')
    def test_fetch_vol_rates_returns_dict(self, mock_dl, provider):
        tickers = list(provider.VOL_RATES.keys())
        mock_dl.return_value = self._patch_download(provider, tickers)
        result = provider.fetch_vol_and_rates()
        assert isinstance(result, dict)

    @patch('yfinance.download')
    def test_fetch_vol_rates_has_vix(self, mock_dl, provider):
        tickers = list(provider.VOL_RATES.keys())
        mock_dl.return_value = self._patch_download(provider, tickers)
        result = provider.fetch_vol_and_rates()
        assert '^VIX' in result

    @patch('yfinance.download')
    def test_fetch_crypto_returns_df(self, mock_dl, provider):
        tickers = list(provider.CRYPTO.keys())
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_crypto()
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.download')
    def test_fetch_crypto_columns(self, mock_dl, provider):
        tickers = list(provider.CRYPTO.keys())
        mock_dl.return_value = self._patch_download(provider, tickers)
        df = provider.fetch_crypto()
        if not df.empty:
            for col in ('name', 'icon', 'last', 'chg_1d', 'chg_7d'):
                assert col in df.columns

    def test_fetch_economic_calendar_returns_df(self, provider):
        df = provider.fetch_economic_calendar()
        assert isinstance(df, pd.DataFrame)
        assert not df.empty

    def test_fetch_economic_calendar_columns(self, provider):
        df = provider.fetch_economic_calendar()
        for col in ('date', 'time', 'flag', 'event', 'importance'):
            assert col in df.columns

    def test_fetch_economic_calendar_has_high_events(self, provider):
        df = provider.fetch_economic_calendar()
        assert 'High' in df['importance'].values

    @patch('yfinance.Ticker')
    def test_fetch_news_returns_df(self, mock_ticker, provider):
        instance = MagicMock()
        instance.news = _fake_ticker_news()
        mock_ticker.return_value = instance
        df = provider.fetch_news(n=5)
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.Ticker')
    def test_fetch_news_columns(self, mock_ticker, provider):
        instance = MagicMock()
        instance.news = _fake_ticker_news()
        mock_ticker.return_value = instance
        df = provider.fetch_news()
        if not df.empty:
            for col in ('headline', 'source', 'ago'):
                assert col in df.columns

    @patch('yfinance.Ticker')
    def test_fetch_news_empty_graceful(self, mock_ticker, provider):
        instance = MagicMock()
        instance.news = []
        mock_ticker.return_value = instance
        df = provider.fetch_news()
        assert isinstance(df, pd.DataFrame)

    @patch('yfinance.download')
    def test_batch_download_empty_returns_df(self, mock_dl, provider):
        mock_dl.return_value = pd.DataFrame()
        result = provider._batch_download(['^GSPC'], period='35d')
        assert isinstance(result, pd.DataFrame)

    @patch('yfinance.download')
    def test_batch_download_exception_returns_df(self, mock_dl, provider):
        mock_dl.side_effect = Exception("network error")
        result = provider._batch_download(['^GSPC'], period='35d')
        assert isinstance(result, pd.DataFrame)
        assert result.empty

    @patch('yfinance.download')
    def test_single_bad_ticker_skipped(self, mock_dl, provider):
        """If one ticker has no data, it should be skipped gracefully."""
        good = ['^GSPC']
        bad  = ['INVALID123']
        tickers = good + bad
        mock_df = _fake_closes(tickers, n=35)
        mock_df_mi = mock_df.copy()
        mock_df_mi.columns = pd.MultiIndex.from_tuples([('Close', t) for t in tickers])
        # Make INVALID123 all NaN
        mock_df_mi[('Close', 'INVALID123')] = np.nan
        mock_dl.return_value = mock_df_mi
        result = provider._batch_download(tickers)
        assert isinstance(result, pd.DataFrame)

    @patch('yfinance.download')
    def test_row_from_closes_change_pct(self, mock_dl, provider):
        s = pd.Series([100.0, 101.0, 99.0, 102.0],
                      index=pd.date_range('2026-03-08', periods=4))
        row = provider._row_from_closes(s, 'Test', '🇺🇸', 'Americas')
        assert row is not None
        assert isinstance(row['change_pct'], float)
        # last=102, prev=99 -> chg = (102-99)/99*100
        assert abs(row['change_pct'] - (102-99)/99*100) < 0.01


# ─────────────────────────────────────────────────────────────────
# HTML Builders
# ─────────────────────────────────────────────────────────────────

class TestHTMLBuilders:
    def test_build_indices_html_not_empty(self):
        from macro_data import _build_indices_html
        df = pd.DataFrame([
            {'name': 'S&P 500', 'flag': '🇺🇸', 'group': 'Americas',
             'last': 6750.0, 'change_pct': 0.5, 'change_net': 30.0, 'sparkline': [100]*20},
        ])
        html = _build_indices_html(df)
        assert 'S&P 500' in html
        assert '<table' in html

    def test_build_indices_html_empty(self):
        from macro_data import _build_indices_html
        html = _build_indices_html(pd.DataFrame())
        assert isinstance(html, str)

    def test_build_fx_html(self):
        from macro_data import _build_fx_html
        df = pd.DataFrame([
            {'name': 'EUR/USD', 'flag': '🇪🇺', 'group': 'G10',
             'last': 1.1556, 'change_pct': -0.3, 'sparkline': [1.15]*25},
        ])
        html = _build_fx_html(df, 'G10')
        assert 'EUR/USD' in html

    def test_build_commodities_html(self):
        from macro_data import _build_commodities_html
        df = pd.DataFrame([
            {'name': 'Gold', 'flag': '🥇', 'group': 'Metals',
             'last': 5182.0, 'change_pct': 0.31, 'sparkline': [5000]*25},
        ])
        html = _build_commodities_html(df)
        assert 'Gold' in html

    def test_build_calendar_html(self):
        from macro_data import _build_calendar_html
        df = pd.DataFrame([
            {'date': 'Mar 12', 'time': '13:30', 'flag': '🇺🇸', 'country': 'US',
             'event': 'CPI', 'importance': 'High', 'previous': '2.8%',
             'forecast': '2.9%', 'actual': ''},
        ])
        html = _build_calendar_html(df)
        assert 'CPI' in html

    def test_build_news_html(self):
        from macro_data import _build_news_html
        df = pd.DataFrame([
            {'ago': '2h ago', 'source': 'Reuters',
             'headline': 'Fed holds rates', 'url': '#',
             'ts': datetime.now(timezone.utc)},
        ])
        html = _build_news_html(df)
        assert 'Fed holds rates' in html

    def test_build_vol_html_with_vix(self):
        from macro_data import _build_vol_html
        vr = {
            '^VIX': {'name': 'VIX', 'cat': 'Vol', 'last': 25.5,
                     'chg_pct': 5.1, 'chg_bps': 0, 'context': 'Elevated — Risk-Off',
                     'sparkline': []},
            '^TNX': {'name': 'US 10Y', 'cat': 'Rates', 'last': 4.208,
                     'chg_pct': 0.5, 'chg_bps': 1.74, 'context': '', 'sparkline': []},
            'curve_slope': 0.5, 'curve_inverted': False,
        }
        html = _build_vol_html(vr)
        assert '25.50' in html
        assert 'VIX' in html

    def test_build_crypto_html(self):
        from macro_data import _build_crypto_html
        df = pd.DataFrame([
            {'name': 'Bitcoin', 'icon': '₿', 'last': 85000.0,
             'chg_1d': 1.5, 'chg_7d': -3.2, 'sparkline': [80000]*20},
        ])
        html = _build_crypto_html(df)
        assert 'Bitcoin' in html


# ─────────────────────────────────────────────────────────────────
# MacroExcelExporter
# ─────────────────────────────────────────────────────────────────

class TestMacroExcelExporter:
    @pytest.fixture
    def sample_data(self):
        return {
            'timestamp': datetime.now(timezone.utc),
            'indices': pd.DataFrame([{
                'name': 'S&P 500', 'flag': '🇺🇸', 'group': 'Americas',
                'last': 6750.0, 'change_pct': 0.5, 'ytd_pct': 3.2,
                'sparkline': [100]*30,
            }]),
            'fx': pd.DataFrame([{
                'name': 'EUR/USD', 'flag': '🇪🇺', 'group': 'G10',
                'last': 1.1556, 'change_pct': -0.3, 'sparkline': [1.15]*30,
            }]),
            'commodities': pd.DataFrame([{
                'name': 'Gold', 'flag': '🥇', 'group': 'Metals',
                'last': 5182.0, 'change_pct': 0.31, 'sparkline': [5000]*30,
            }]),
            'bonds': pd.DataFrame([{
                'name': 'US 10Y', 'flag': '🇺🇸', 'maturity': 10,
                'yield_pct': 4.208, 'chg_bps': 1.74, 'spread_vs_us': 0.0,
                'sparkline': [4.0]*30,
            }]),
            'vol_rates': {
                '^VIX': {'name': 'VIX', 'cat': 'Vol', 'last': 25.5,
                         'chg_pct': 5.1, 'chg_bps': 0, 'context': ''},
                'curve_slope': 0.5, 'curve_inverted': False,
            },
            'crypto': pd.DataFrame([{
                'name': 'Bitcoin', 'icon': '₿', 'last': 85000.0,
                'chg_1d': 1.5, 'chg_7d': -3.2, 'sparkline': [80000]*30,
            }]),
            'calendar': pd.DataFrame([{
                'date': 'Mar 12', 'time': '13:30', 'flag': '🇺🇸', 'country': 'US',
                'event': 'CPI', 'importance': 'High',
                'previous': '2.8%', 'forecast': '2.9%', 'actual': '',
            }]),
            'news': pd.DataFrame(),
            'errors': [],
        }

    def test_export_bytes_returns_bytes(self, sample_data):
        from macro_export import MacroExcelExporter
        exp = MacroExcelExporter()
        result = exp.export_bytes(sample_data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_bytes_is_valid_xlsx(self, sample_data):
        from macro_export import MacroExcelExporter
        import openpyxl
        exp = MacroExcelExporter()
        xl_bytes = exp.export_bytes(sample_data)
        # Should be readable by openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
        assert 'Dashboard' in wb.sheetnames

    def test_export_bytes_has_equity_sheet(self, sample_data):
        from macro_export import MacroExcelExporter
        import openpyxl
        exp = MacroExcelExporter()
        wb = openpyxl.load_workbook(io.BytesIO(exp.export_bytes(sample_data)))
        assert 'Equity Indices' in wb.sheetnames

    def test_export_bytes_with_empty_data(self):
        from macro_export import MacroExcelExporter
        exp = MacroExcelExporter()
        data = {'timestamp': datetime.now(timezone.utc), 'errors': []}
        result = exp.export_bytes(data)
        assert isinstance(result, bytes)


# ─────────────────────────────────────────────────────────────────
# MacroEmailSender
# ─────────────────────────────────────────────────────────────────

class TestMacroEmailSender:
    @pytest.fixture
    def sample_data(self):
        return {
            'timestamp': datetime.now(timezone.utc),
            'indices': pd.DataFrame([{
                'name': 'S&P 500', 'flag': '🇺🇸', 'group': 'Americas',
                'last': 6750.0, 'change_pct': 0.5,
            }]),
            'fx': pd.DataFrame([{
                'name': 'EUR/USD', 'flag': '🇪🇺', 'group': 'G10',
                'last': 1.1556, 'change_pct': -0.3,
            }]),
            'commodities': pd.DataFrame(),
            'vol_rates': {'^VIX': {'name': 'VIX', 'last': 25.5, 'chg_pct': 5.1}},
            'calendar': pd.DataFrame([{
                'date': 'Mar 12', 'time': '13:30', 'flag': '🇺🇸',
                'event': 'CPI', 'importance': 'High',
                'previous': '', 'forecast': '', 'actual': '',
            }]),
            'crypto': pd.DataFrame(),
            'news': pd.DataFrame(),
            'errors': [],
        }

    def test_build_html_body_returns_string(self, sample_data):
        from macro_export import MacroEmailSender
        sender = MacroEmailSender()
        html = sender.build_html_body(sample_data)
        assert isinstance(html, str)
        assert len(html) > 100

    def test_build_html_body_contains_ravinala(self, sample_data):
        from macro_export import MacroEmailSender
        html = MacroEmailSender().build_html_body(sample_data)
        assert 'RAVINALA' in html

    def test_build_html_body_contains_timestamp(self, sample_data):
        from macro_export import MacroEmailSender
        html = MacroEmailSender().build_html_body(sample_data)
        # Should contain year
        assert '2026' in html

    def test_build_html_body_contains_index(self, sample_data):
        from macro_export import MacroEmailSender
        html = MacroEmailSender().build_html_body(sample_data)
        assert 'S&P 500' in html

    def test_send_fails_without_credentials(self, sample_data):
        from macro_export import MacroEmailSender
        sender = MacroEmailSender(smtp_user='', smtp_password='')
        result = sender.send_snapshot('test@example.com', sample_data)
        assert result is False

    def test_send_smtp_error_returns_false(self, sample_data):
        from macro_export import MacroEmailSender
        sender = MacroEmailSender(smtp_user='u', smtp_password='p')
        # Should return False on connection error (no real SMTP)
        result = sender.send_snapshot('test@example.com', sample_data, attach_excel=False)
        assert result is False
