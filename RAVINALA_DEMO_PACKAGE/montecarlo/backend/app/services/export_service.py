"""
services/export_service.py - shared dashboard export service.

Owns snapshot retrieval, file assembly, filename/path generation and export
metadata so HTTP routes stay as thin controllers.
"""

from __future__ import annotations

import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from app.models import ExcelExportRequest, PDFExportRequest
from app.services.snapshot_service import get_full_snapshot_async


@dataclass(frozen=True)
class ExportedFile:
    """Resolved export artifact returned by the service layer."""

    path: str
    filename: str
    media_type: str


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _export_dir() -> Path:
    path = Path(tempfile.gettempdir()) / "ravinala_exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _timestamp_slug() -> str:
    return _utcnow().strftime("%Y%m%d_%H%M%S")


def _write_excel(snapshot: dict[str, Any], request: ExcelExportRequest) -> ExportedFile:
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    if "indices" in request.sheets:
        worksheet = workbook.create_sheet("Indices")
        worksheet.append(["Symbol", "Name", "Region", "Price", "Change %", "Updated"])
        for _, items in snapshot.get("indices", {}).items():
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        worksheet.append([
                            item.get("symbol", ""),
                            item.get("name", ""),
                            item.get("region", ""),
                            item.get("price", 0),
                            item.get("change", {}).get("percent", 0),
                            item.get("timestamp", ""),
                        ])

    if "bonds" in request.sheets:
        worksheet = workbook.create_sheet("Bonds")
        worksheet.append(["Country", "2Y Yield", "5Y Yield", "10Y Yield", "Spread vs Bund"])
        for bond in snapshot.get("bonds", {}).get("bonds", []):
            worksheet.append([
                bond.get("country", ""),
                bond.get("yield_2y", ""),
                bond.get("yield_5y", ""),
                bond.get("yield_10y", ""),
                bond.get("spread_vs_bund_bp", ""),
            ])

    filename = f"ravinala_dashboard_{_timestamp_slug()}.xlsx"
    path = _export_dir() / filename
    workbook.save(path)
    return ExportedFile(
        path=str(path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _write_pdf(snapshot: dict[str, Any], request: PDFExportRequest) -> ExportedFile:
    filename = f"ravinala_dashboard_{_timestamp_slug()}.pdf"
    path = _export_dir() / filename
    generated_at = _utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    pdf = canvas.Canvas(str(path), pagesize=landscape(letter))
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(1 * inch, 10 * inch, "Ravinala Global Macro Dashboard")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(1 * inch, 9.5 * inch, f"Generated: {generated_at}")
    pdf.drawString(1 * inch, 9.1 * inch, f"Layout: {request.layout}")
    pdf.drawString(1 * inch, 8.7 * inch, f"Sections: {', '.join(sorted(snapshot.keys()))}")
    pdf.save()

    return ExportedFile(
        path=str(path),
        filename=filename,
        media_type="application/pdf",
    )


async def export_dashboard_excel(request: ExcelExportRequest) -> ExportedFile:
    """Generate an Excel dashboard export file from the current snapshot."""
    snapshot = await get_full_snapshot_async()
    return _write_excel(snapshot, request)


async def export_dashboard_pdf(request: PDFExportRequest) -> ExportedFile:
    """Generate a PDF dashboard export file from the current snapshot."""
    snapshot = await get_full_snapshot_async()
    return _write_pdf(snapshot, request)


__all__ = [
    "ExportedFile",
    "export_dashboard_excel",
    "export_dashboard_pdf",
]
